"""Data export manager — exports Kahunas data to Excel with organized directory structure."""

from __future__ import annotations

import logging
import re
from datetime import UTC, datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from ..client import KahunasClient

logger = logging.getLogger(__name__)

_HEADER_FONT = Font(bold=True, size=12)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _sanitize_name(name: str) -> str:
    """Make a filesystem-safe name."""
    return re.sub(r'[<>:"/\\|?*]', "_", name).strip().rstrip(".")


def _timestamp() -> str:
    return datetime.now(tz=UTC).strftime("%Y%m%d_%H%M%S")


def _add_header_row(ws: Any, headers: list[str]) -> None:
    """Add a styled header row to a worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: Any) -> None:
    """Auto-fit column widths."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, min(len(val), 60))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)


class ExportManager:
    """Exports Kahunas data to user-friendly Excel files."""

    def __init__(self, client: KahunasClient) -> None:
        self._client = client

    def _base_dir(self, output_dir: str | None) -> Path:
        if output_dir:
            return Path(output_dir)
        return Path.home() / "kahunas_exports" / _timestamp()

    async def export_client(
        self,
        client_uuid: str,
        output_dir: str | None = None,
        include_photos: bool = True,
        include_checkins: bool = True,
        include_progress: bool = True,
        include_workouts: bool = True,
        include_habits: bool = True,
        include_chat: bool = True,
    ) -> Path:
        """Export all data for a single client."""
        # Get client info
        resp = await self._client.get_client_action("view", client_uuid)
        client_data = self._parse_response(resp)

        client_name = self._extract_client_name(client_data, client_uuid)
        base = self._base_dir(output_dir)
        client_dir = base / _sanitize_name(client_name)
        client_dir.mkdir(parents=True, exist_ok=True)

        # Profile
        self._export_client_profile(client_dir, client_data)

        # Check-ins
        if include_checkins:
            await self._export_client_checkins(client_dir, client_uuid, include_photos)

        # Progress
        if include_progress:
            await self._export_client_progress(client_dir, client_uuid)

        # Habits
        if include_habits:
            await self._export_client_habits(client_dir, client_uuid)

        # Chat
        if include_chat:
            await self._export_client_chat(client_dir, client_uuid)

        logger.info("Exported client %s to %s", client_name, client_dir)
        return client_dir

    async def export_all_clients(self, output_dir: str | None = None) -> Path:
        """Export data for all clients."""
        base = self._base_dir(output_dir)
        base.mkdir(parents=True, exist_ok=True)

        resp = await self._client.list_clients()
        clients_data = self._parse_response(resp)

        clients = []
        if isinstance(clients_data, dict):
            clients = clients_data.get("data", clients_data.get("clients", []))
        if isinstance(clients_data, list):
            clients = clients_data

        if not clients:
            # Write empty summary
            wb = Workbook()
            ws = wb.active
            ws.title = "Clients"
            _add_header_row(ws, ["Status"])
            ws.cell(row=2, column=1, value="No clients found")
            wb.save(base / "clients_summary.xlsx")
            return base

        for client in clients:
            uuid = client.get("uuid", client.get("id", ""))
            if uuid:
                try:
                    await self.export_client(str(uuid), output_dir=str(base))
                except Exception as e:
                    logger.warning("Failed to export client %s: %s", uuid, e)

        return base

    async def export_exercise_library(self, output_dir: str | None = None) -> Path:
        """Export the full exercise library to Excel."""
        base = self._base_dir(output_dir)
        base.mkdir(parents=True, exist_ok=True)
        filepath = base / "exercise_library.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Exercises"

        headers = [
            "Name",
            "Type",
            "Sets",
            "Reps",
            "RIR",
            "RPE",
            "Intensity",
            "Rest Period",
            "Tempo",
            "Notes",
            "Tags",
        ]
        _add_header_row(ws, headers)

        page = 1
        row = 2
        while True:
            data = await self._client.list_exercises(page=page, per_page=100)
            for ex in data.exercises:
                ws.cell(row=row, column=1, value=ex.exercise_name or ex.title)
                ws.cell(row=row, column=2, value="Strength" if ex.exercise_type == 1 else "Cardio")
                ws.cell(row=row, column=3, value=ex.sets)
                ws.cell(row=row, column=4, value=ex.reps)
                ws.cell(row=row, column=5, value=ex.rir)
                ws.cell(row=row, column=6, value=ex.rpe_rating)
                ws.cell(row=row, column=7, value=ex.intensity)
                ws.cell(row=row, column=8, value=ex.rest_period)
                ws.cell(row=row, column=9, value=ex.tempo)
                ws.cell(row=row, column=10, value=ex.notes)
                ws.cell(row=row, column=11, value=", ".join(ex.tags))
                row += 1

            if not data.pagination.next_page:
                break
            page += 1

        _auto_width(ws)
        wb.save(filepath)
        logger.info("Exported exercise library to %s", filepath)
        return filepath

    async def export_workout_programs(self, output_dir: str | None = None) -> Path:
        """Export all workout programs to Excel files."""
        base = self._base_dir(output_dir)
        programs_dir = base / "workout_programs"
        programs_dir.mkdir(parents=True, exist_ok=True)

        page = 1
        while True:
            data = await self._client.list_workout_programs(page=page, per_page=100)
            for program_summary in data.workout_plan:
                try:
                    detail_data = await self._client.get_workout_program(program_summary.uuid)
                    program = detail_data.workout_plan
                    self._export_single_program(programs_dir, program)
                except Exception as e:
                    logger.warning("Failed to export program %s: %s", program_summary.uuid, e)

            if not data.pagination.next_page:
                break
            page += 1

        return programs_dir

    def _export_single_program(self, programs_dir: Path, program: Any) -> None:
        """Export a single workout program to an Excel file."""
        name = _sanitize_name(program.title or program.uuid)
        filepath = programs_dir / f"{name}.xlsx"

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        if not program.workout_days:
            ws = wb.create_sheet("Overview")
            _add_header_row(ws, ["Program", "Description"])
            ws.cell(row=2, column=1, value=program.title)
            ws.cell(row=2, column=2, value=program.long_desc or program.short_desc)
            wb.save(filepath)
            return

        for day in program.workout_days:
            sheet_name = _sanitize_name(day.title or "Rest Day")[:31]
            ws = wb.create_sheet(sheet_name)

            if day.is_restday:
                ws.cell(row=1, column=1, value="Rest Day")
                ws["A1"].font = _HEADER_FONT
                continue

            headers = [
                "Exercise",
                "Group Type",
                "Sets",
                "Reps",
                "Weight",
                "RIR",
                "RPE",
                "Tempo",
                "Rest Period",
                "Notes",
            ]
            _add_header_row(ws, headers)

            row = 2
            for section_name in ("warmup", "workout", "cooldown"):
                groups = getattr(day.exercise_list, section_name, [])
                if groups:
                    ws.cell(row=row, column=1, value=f"── {section_name.upper()} ──")
                    ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
                    row += 1

                for group in groups:
                    group_type = group.type if group.type != "normal" else ""
                    for exercise in group.exercises:
                        ws.cell(row=row, column=1, value=exercise.exercise_name)
                        ws.cell(row=row, column=2, value=group_type)
                        ws.cell(row=row, column=3, value=exercise.sets)
                        ws.cell(row=row, column=4, value=exercise.reps)
                        ws.cell(row=row, column=5, value="")
                        ws.cell(row=row, column=6, value=exercise.rir)
                        ws.cell(row=row, column=7, value=exercise.rpe_rating)
                        ws.cell(row=row, column=8, value=exercise.tempo)
                        ws.cell(row=row, column=9, value=exercise.rest_period)
                        ws.cell(row=row, column=10, value=exercise.notes or "")
                        row += 1

            _auto_width(ws)

        wb.save(filepath)

    def _export_client_profile(self, client_dir: Path, client_data: Any) -> None:
        """Export client profile to Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Profile"

        # Row 1: header
        for col, header in enumerate(["Field", "Value"], 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = _HEADER_FONT_WHITE
            cell.fill = _HEADER_FILL

        if isinstance(client_data, dict):
            data = client_data.get("data", client_data)
            if isinstance(data, dict):
                row = 2

                for key, val in data.items():
                    if isinstance(val, (str, int, float, bool)):
                        # Human-readable field names
                        label = key.replace("_", " ").title()
                        ws.cell(row=row, column=1, value=label)
                        ws.cell(row=row, column=2, value=str(val))
                        row += 1

        _auto_width(ws)
        wb.save(client_dir / "profile.xlsx")

    async def _export_client_checkins(
        self, client_dir: Path, client_uuid: str, include_photos: bool
    ) -> None:
        """Export client check-ins to Excel."""
        checkins_dir = client_dir / "checkins"
        checkins_dir.mkdir(exist_ok=True)

        # Try to get check-in list via client view
        resp = await self._client.get_client_action("view", client_uuid)
        data = self._parse_response(resp)

        checkins = []
        if isinstance(data, dict):
            checkins = data.get("checkins", data.get("check_ins", []))

        if not checkins:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Check-ins Summary"
        _add_header_row(ws, ["#", "Date", "UUID", "Status"])

        for i, ci in enumerate(checkins, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=ci.get("check_in_number", i))
            ws.cell(row=row, column=2, value=ci.get("submitted_at", ci.get("date", "")))
            ws.cell(row=row, column=3, value=ci.get("uuid", ""))
            ws.cell(row=row, column=4, value=ci.get("status", "submitted"))

            # Download photos if requested
            if include_photos:
                photos = ci.get("photos", ci.get("images", []))
                if photos:
                    photos_dir = checkins_dir / "photos"
                    photos_dir.mkdir(exist_ok=True)
                    await self._download_photos(photos, photos_dir, prefix=f"checkin_{i}")

        _auto_width(ws)
        wb.save(checkins_dir / "checkins_summary.xlsx")

    async def _export_client_progress(self, client_dir: Path, client_uuid: str) -> None:
        """Export client progress data to Excel."""
        progress_dir = client_dir / "progress"
        progress_dir.mkdir(exist_ok=True)

        metrics = ["weight", "bodyfat", "chest", "waist", "hips", "arms", "thighs"]
        wb = Workbook()
        wb.remove(wb.active)

        for metric in metrics:
            try:
                resp = await self._client.get_chart_data(client_uuid, value=metric)
                data = self._parse_response(resp)
                if not data:
                    continue

                ws = wb.create_sheet(metric.title())
                _add_header_row(ws, ["Date", metric.title()])

                chart_data = data if isinstance(data, list) else data.get("data", [])
                if isinstance(chart_data, list):
                    for i, point in enumerate(chart_data, 2):
                        if isinstance(point, dict):
                            ws.cell(
                                row=i, column=1, value=point.get("date", point.get("label", ""))
                            )
                            ws.cell(row=i, column=2, value=point.get("value", point.get("y", "")))

                _auto_width(ws)
            except Exception as e:
                logger.debug("No %s data for client: %s", metric, e)

        if wb.sheetnames:
            wb.save(progress_dir / "body_measurements.xlsx")

    async def _export_client_habits(self, client_dir: Path, client_uuid: str) -> None:
        """Export client habits to Excel."""
        resp = await self._client.list_habits(client_uuid)
        data = self._parse_response(resp)

        habits = []
        if isinstance(data, dict):
            habits = data.get("habits", data.get("data", []))
        if isinstance(data, list):
            habits = data

        if not habits:
            return

        habits_dir = client_dir / "habits"
        habits_dir.mkdir(exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Habits"
        _add_header_row(ws, ["Habit", "Date", "Completed", "UUID"])

        for i, habit in enumerate(habits, 2):
            ws.cell(row=i, column=1, value=habit.get("title", ""))
            ws.cell(row=i, column=2, value=habit.get("date", ""))
            ws.cell(row=i, column=3, value="Yes" if habit.get("completed") else "No")
            ws.cell(row=i, column=4, value=habit.get("uuid", ""))

        _auto_width(ws)
        wb.save(habits_dir / "habit_tracking.xlsx")

    async def _export_client_chat(self, client_dir: Path, client_uuid: str) -> None:
        """Export chat messages to Excel."""
        resp = await self._client.get_chat_messages(client_uuid)
        data = self._parse_response(resp)

        messages = []
        if isinstance(data, dict):
            messages = data.get("messages", data.get("data", []))
        if isinstance(data, list):
            messages = data

        if not messages:
            return

        chat_dir = client_dir / "chat"
        chat_dir.mkdir(exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        _add_header_row(ws, ["Date", "From", "Message", "Read"])

        for i, msg in enumerate(messages, 2):
            ws.cell(row=i, column=1, value=msg.get("created_at", ""))
            sender = msg.get("sender_name", msg.get("sender_uuid", ""))
            ws.cell(row=i, column=2, value=sender)
            ws.cell(row=i, column=3, value=msg.get("message", ""))
            ws.cell(row=i, column=3).alignment = _WRAP
            ws.cell(row=i, column=4, value="Yes" if msg.get("read") else "No")

        _auto_width(ws)
        wb.save(chat_dir / "chat_history.xlsx")

    async def _download_photos(self, photos: list[Any], target_dir: Path, prefix: str = "") -> None:
        """Download photos to a directory."""
        async with httpx.AsyncClient(timeout=30) as http:
            for i, photo in enumerate(photos):
                url = ""
                if isinstance(photo, str):
                    url = photo
                elif isinstance(photo, dict):
                    url = photo.get("file_url", photo.get("url", photo.get("image_url", "")))

                if not url:
                    continue

                try:
                    resp = await http.get(url)
                    if resp.status_code == 200:
                        ext = url.rsplit(".", 1)[-1][:4] if "." in url else "jpg"
                        filename = f"{prefix}_{i + 1}.{ext}" if prefix else f"photo_{i + 1}.{ext}"
                        (target_dir / filename).write_bytes(resp.content)
                except Exception as e:
                    logger.debug("Failed to download photo %s: %s", url, e)

    @staticmethod
    def _parse_response(resp: httpx.Response) -> Any:
        """Parse an httpx response to dict/list."""
        try:
            return resp.json()
        except Exception:
            return {}

    @staticmethod
    def _extract_client_name(data: Any, fallback: str) -> str:
        """Extract a readable client name from response data."""
        if isinstance(data, dict):
            d = data.get("data", data)
            if isinstance(d, dict):
                first = d.get("first_name", "")
                last = d.get("last_name", "")
                if first or last:
                    return f"{first} {last}".strip()
                return d.get("name", d.get("email", fallback))
        return fallback
